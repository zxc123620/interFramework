import logging

import allure
import openpyxl
import pytest

from config import ROOT_PATH
from utils.api_keywords.ApiKeyWords import ApiKeyword
from utils.assert_utils.AssertType import AssertType, AssertMethod
from utils.excel_utils.read_excel import read_excel
import utils.log_utils.log_config  # 不能删除
from utils.my_exception.all_exception import AssertTypeException, AssertNumNotMatchException
from config import EXCEL_FILE_PATH, TEST_SHEET_NAME

logger = logging.getLogger("main.create_test")
ak = ApiKeyword()
excel = openpyxl.load_workbook(EXCEL_FILE_PATH)
sheet = excel[TEST_SHEET_NAME]
all_val = {}
result = ""


@pytest.mark.parametrize('data', read_excel())
def test_single(data):
    # # 动态生成标题
    # allure.dynamic.title(data[11])
    # 如果存在自定义标题
    logger.info("======================================================")
    logger.info("用例标题: %s" % data['用例名'])
    logger.info("数据: %s" % data)
    global result
    if data['用例名'] is not None:
        # 动态生成标题
        allure.dynamic.title(data['用例名'])
    if data['story(小模块)'] is not None:
        # 动态获取story模块名
        allure.dynamic.story(data['story(小模块)'])
    if data['feature(大模块)'] is not None:
        # 动态获取feature模块名
        allure.dynamic.feature(data['feature(大模块)'])
    if data['备注'] is not None:
        # 动态获取备注信息
        allure.dynamic.description(data['备注'])
    if data['级别'] is not None:
        # 动态获取级别信息(blocker、critical、normal、minor、trivial)
        allure.dynamic.severity('级别')
    # 行数
    r = str(data["编号"] + 1)
    # 解析文件
    file_str = data["文件"]
    if file_str is None:
        files = None
    else:
        file_list = []
        file_dict = eval(file_str)
        for key, value in file_dict.items():
            value = ROOT_PATH + "/upload_files/" + value
            file_list.append((key, (key, open(value, "rb"))))  # ('1.png', ('1.png', open('logo.png', 'rb')))
        files = tuple(file_list)
    # ==============Excel数据解析==============
    try:
        url = data["地址"] + data["路径"] if data["路径"] is not None else data["地址"]
        dict_data = {
            'url': url,
            'params': eval(data['公共参数(PARAMS)']),
            'headers': eval(data['请求头']),
            data["参数类型"]: eval(data['参数']),  # json/data
            "files": files
        }
    except Exception:
        logger.error("接口请求格式有误，请检查url、params、headers、data、参数类型、文件")
        sheet["N" + r] = "请求参数有误，请检查"
        raise
    res = getattr(ak, data['请求方法'])(**dict_data)
    # =================Json提取器=================
    if data['JSON提取_引用名称'] is not None:
        # 遍历分割JSON提取_引用名称
        json_name_str = data['JSON提取_引用名称']
        # 用分号分割varStr字符串，并保存到列表
        json_name_str_list = json_name_str.split(';')
        logger.info("JSON提取_引用名称: %s" % json_name_str_list)
        # 获取列表长度
        length = len(json_name_str_list)
        # 遍历分割JSON表达式
        jsonpath_str = data['JSON表达式']
        jsonpath_str_list = jsonpath_str.split(';')
        logger.info("JSON表达式: %s" % jsonpath_str_list)
        # 循环输出列表值
        for i in range(length):
            # 获取JSON提取_引用名称
            key = json_name_str_list[i]
            # json表达式获取
            json_exp = jsonpath_str_list[i]
            # 字典值获取
            value_json = ak.get_text(res.text, json_exp)
            # 持续添加参数，只要参数名不重复，重复的后面就会覆盖前面的参数
            all_val[key] = value_json
    # =================校验类型校验=================
    # 校验类型校验
    assert_types = [i.value for i in AssertType]
    excel_assert_types = str(data["校验类型"]).split(";")  # 校验类型(多个)
    excel_json_assert_columns = data['校验字段'].split(";")  # 校验字段(多个)
    excel_expected_values = data['预期结果'].split(";")  # 预期结果(多个)
    if not (len(excel_assert_types) == len(excel_json_assert_columns) == len(excel_expected_values)):  # 查看数量是否一致
        info = "校验类型、字段、预期结果数量不匹配: %s %s %s" % (len(excel_assert_types), len(excel_json_assert_columns), len(
            excel_expected_values))
        logger.error(info)
        raise AssertNumNotMatchException(info)
    # =================预期结果进行转换操作=================
    # 预期结果进行转换操作
    temp_list = []
    for i in excel_expected_values:
        if i.startswith("{") or i.startswith("["):
            temp_list.append(eval(i))
        else:
            temp_list.append(i)
    excel_expected_values = temp_list

    # 数量相等的情况
    # =================校验操作=================
    for i in range(len(excel_assert_types)):  # 校验参数类型
        excel_assert_type = excel_assert_types[i]  # 校验类型(单个)
        excel_json_assert_column = excel_json_assert_columns[i]  # json校验字段(单个)
        excel_expected_value = excel_expected_values[i]  # 预期结果(单个)
        if excel_assert_type not in assert_types:  # 校验类型判断
            error_info = "校验类型[ %s ]不正确,可用校验参数: %s " % (excel_assert_type, assert_types)
            logger.error(error_info)
            raise AssertTypeException(error_info)
        # 参数类型校验成功
        excel_assert_name = AssertType(excel_assert_type).name
        # 结果校验
        logger.info("=====进行校验 %s" % i)
        result = ak.get_text(res.text, excel_json_assert_column, multiple=isinstance(excel_expected_value, (list, dict)))
        logger.info("预期结果: %s" % excel_expected_value)
        logger.info("实际结果(false表示没找到): %s" % result)
        is_ok, msg = getattr(AssertMethod, excel_assert_name)(excel_expected_value, result)
        sheet["N" + r] = msg
        assert is_ok, msg
    excel.save(EXCEL_FILE_PATH)

    # 数据库结果校验
    # 如果存在数据库检查并且需要动态关联接口参数
    # if data['数据库变量'] is not None:
    #     try:
    #         # 拼接sql
    #         str_sql = data['数据库SQL']
    #         # 遍历分割数据库变量
    #         sql_param_str = data['数据库变量']
    #         sql_param_list = sql_param_str.split(';')  # ["all_val['VAR_UID']", "all_val['VAR_UNAME']"]
    #         list1 = []
    #         # 获取sql参数列表长度
    #         length = len(sql_param_list)
    #         # 循环解析all_val[]值并填入list1
    #         for i in range(length):
    #             list1.append(eval(sql_param_list[i]))
    #         sql = str_sql.format(*list1)
    #         # 实际结果
    #         with MysqlPool as db:
    #             db.cursor.execute(sql)
    #             sql_check = db.cursor.fetchone().values
    #             if sql_check == data[22]:
    #                 sheet.cell(r, 11).value = "通过"
    #             else:
    #                 sheet.cell(r, 11).value = "不通过"
    #             excel.save(excel_path)
    #     except:
    #         print("=============实际结果=================")
    #         print("sql有误，请检查")
    #         sheet.cell(r, 11).value = "sql有误，请检查"
    #         excel.save(excel_path)
    #     finally:
    #         # assert_utils sql_check == data[22]
    #         pass
    # # 数据库结果校验
    # # 如果存在数据库检查，不需要动态关联接口参数
    # if data[20] is not None:
    #     try:
    #         # 实际结果
    #         sql = None
    #         sql_check = None
    #         sql = data[20]
    #         sql_check = ak.sqlCheck(sql, 1)
    #         if (sql_check == data[22]):
    #             sheet.cell(r, 11).value = "通过"
    #         else:
    #             sheet.cell(r, 11).value = "不通过"
    #         excel.save(excel_path)
    #     except:
    #         print("=============实际结果=================")
    #         print("sql有误，请检查")
    #         sheet.cell(r, 11).value = "sql有误，请检查"
    #         excel.save(excel_path)
    #     finally:
    #         assert_utils sql_check == data[22]
